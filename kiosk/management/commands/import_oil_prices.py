"""
Excel 단가표에서 차종별 오일 가격을 임포트하는 커맨드.

사용법:
    python manage.py import_oil_prices data/퀵오일_차종별_오일별_단가표_260110.xlsx
    python manage.py import_oil_prices data/퀵오일_차종별_오일별_단가표_260110.xlsx --dry-run
    python manage.py import_oil_prices data/퀵오일_차종별_오일별_단가표_260110.xlsx --clear
"""
import re
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from kiosk.models import CarBrand, CarModel, FuelType, OilProduct, OilPrice


# 시트명(부분매칭) → OilProduct tier
SHEET_TIER_MAP = [
    ('킥스 GX5', 'economy'),
    ('킥스 GX7', 'standard'),
    ('킥스 Pao', 'premium'),
    ('벤졸', 'premium_hybrid'),
    ('슈퍼노멀', 'hyperformance'),
    ('메탈로센', 'racing'),
]

# 시트 내 브랜드 그룹별 열 배치 (0-indexed column)
# (brand_name, car_col, gasoline_col, diesel_col_or_none)
BRAND_COLUMNS = [
    ('현대', 1, 2, 3),
    ('기아', 5, 6, 7),
    ('제네시스', 9, 10, None),
    ('르노코리아', 12, 13, 14),
]

# 엑셀에서 "쏘나타 DN8" 형식으로 세대 분리되는 차종
GENERATION_PARENTS = {'쏘나타', '그랜져', '아반떼'}

# 세대명 없이 부모명만 등장할 때 기본 세대명 매핑
DEFAULT_GENERATION = {
    '아반떼': 'CN7',
}

# 헤더/브랜드 키워드 (무시할 셀값)
SKIP_VALUES = frozenset({
    '차종', '현대', '기아', '제네시스', '르노', '르노코리아',
    '휘발유', '경유', '휘발유/LPG', '휘발유/LPG/하이',
    '하이브리드',
})

# 서브 브랜드 감지 키워드
SUB_BRAND_DETECT = {
    'KGM': 'KG모빌리티',
    'KG모빌리티': 'KG모빌리티',
    '쉐보레': '쉐보레',
}


def match_sheet_tier(sheet_name):
    for keyword, tier in SHEET_TIER_MAP:
        if keyword in sheet_name:
            return tier
    return None


def parse_car_name(raw_name):
    """차종명 파싱 → (모델명, 세대명 or None)"""
    if not raw_name:
        return None, None
    raw_name = str(raw_name).strip()
    if not raw_name:
        return None, None

    # 공백으로 분리: "쏘나타 DN8" → ("쏘나타", "DN8")
    parts = raw_name.split(None, 1)
    if len(parts) == 2 and parts[0] in GENERATION_PARENTS:
        return parts[0], parts[1]

    # "아반떼" → ("아반떼", "CN7") via DEFAULT_GENERATION
    if raw_name in DEFAULT_GENERATION:
        return raw_name, DEFAULT_GENERATION[raw_name]

    return raw_name, None


def parse_price(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = int(value)
        return v if v > 0 else None
    if isinstance(value, str):
        value = value.strip()
        if value in ('-', '', '—', '–'):
            return None
        cleaned = re.sub(r'[^\d]', '', value)
        return int(cleaned) if cleaned else None
    return None


def is_sub_brand_header(cell_value):
    """서브 브랜드 헤더인지 확인. 매칭되면 brand_name 반환."""
    if not cell_value:
        return None
    val = str(cell_value).strip()
    # "KGM(쌍 용)" → starts with KGM
    for keyword, brand_name in SUB_BRAND_DETECT.items():
        if val == keyword or val.startswith(keyword + '('):
            return brand_name
    return None


class Command(BaseCommand):
    help = 'Excel 단가표에서 차종별 오일 가격을 임포트합니다.'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Excel 파일 경로')
        parser.add_argument('--dry-run', action='store_true', help='실제 DB 변경 없이 결과만 출력')
        parser.add_argument('--clear', action='store_true', help='기존 가격 데이터 삭제 후 재임포트')

    def handle(self, *args, **options):
        filepath = options['file']
        dry_run = options['dry_run']
        clear = options['clear']

        if dry_run:
            self.stdout.write(self.style.WARNING('=== DRY RUN 모드 ==='))

        wb = load_workbook(filepath, data_only=True)

        # 캐시 로드
        self.fuel_types = {ft.name: ft for ft in FuelType.objects.all()}
        self.oil_products = {op.tier: op for op in OilProduct.objects.all()}
        self.brand_cache = {b.name: b for b in CarBrand.objects.all()}

        self.model_cache = {}
        for m in CarModel.objects.select_related('parent').all():
            key = (m.brand_id, m.name, m.parent_id)
            self.model_cache[key] = m

        self.stats = {
            'models_created': 0,
            'prices_created': 0,
            'prices_updated': 0,
            'prices_cleared': 0,
            'skipped': 0,
        }

        # --clear: 기존 가격 삭제
        if clear and not dry_run:
            count = OilPrice.objects.count()
            OilPrice.objects.all().delete()
            self.stats['prices_cleared'] = count
            self.stdout.write(self.style.WARNING(f'기존 가격 {count}건 삭제'))

        # 아반떼 CN7 세대 보정
        if not dry_run:
            self._ensure_avante_cn7()

        # 시트 처리
        for sheet_name in wb.sheetnames:
            tier = match_sheet_tier(sheet_name)
            if tier is None:
                continue

            oil_product = self.oil_products.get(tier)
            if not oil_product:
                self.stdout.write(self.style.ERROR(f'OilProduct 없음: tier={tier}'))
                continue

            self.stdout.write(f'\n처리: {sheet_name} → {tier}')
            ws = wb[sheet_name]
            self._process_sheet(ws, oil_product, tier, dry_run)

        # 결과
        self.stdout.write(self.style.SUCCESS('\n=== 결과 ==='))
        for k, v in self.stats.items():
            self.stdout.write(f'  {k}: {v}')

    def _ensure_avante_cn7(self):
        """아반떼 CN7 세대 보정: 부모에 직접 걸린 가격을 CN7 세대로 이동"""
        brand = self.brand_cache.get('현대')
        if not brand:
            return

        parent_key = (brand.id, '아반떼', None)
        parent = self.model_cache.get(parent_key)
        if not parent:
            return

        # CN7 세대 생성
        gen_key = (brand.id, 'CN7', parent.id)
        if gen_key not in self.model_cache:
            gen, created = CarModel.objects.get_or_create(
                brand=brand, name='CN7', parent=parent,
                defaults={'order': 0},
            )
            self.model_cache[gen_key] = gen
            if created:
                self.stats['models_created'] += 1
                self.stdout.write('아반떼 CN7 세대 생성')

        cn7 = self.model_cache[gen_key]

        # 부모 가격 → CN7로 이동
        parent_prices = OilPrice.objects.filter(car_model=parent)
        moved = 0
        for p in parent_prices:
            OilPrice.objects.update_or_create(
                car_model=cn7, oil_product=p.oil_product, fuel_type=p.fuel_type,
                defaults={'price': p.price},
            )
            moved += 1
        if moved:
            parent_prices.delete()
            self.stdout.write(f'아반떼 부모 가격 {moved}건 → CN7 이동')

    def _process_sheet(self, ws, oil_product, tier, dry_run):
        """시트 하나 처리 - 4개 브랜드 블록 + 서브 브랜드"""
        for brand_name, car_col, gas_col, diesel_col in BRAND_COLUMNS:
            self._process_block(
                ws, oil_product, tier, brand_name,
                car_col, gas_col, diesel_col, dry_run,
            )

    def _process_block(self, ws, oil_product, tier, brand_name,
                       car_col, gas_col, diesel_col, dry_run,
                       start_row=5, end_row=None):
        """브랜드 블록 하나 처리"""
        if end_row is None:
            end_row = ws.max_row

        brand = self._get_or_create_brand(brand_name, dry_run)
        count = 0
        sub_brand_found = False

        for row in range(start_row, end_row + 1):
            raw = ws.cell(row=row, column=car_col + 1).value  # openpyxl 1-indexed
            if not raw:
                continue
            raw = str(raw).strip()
            if not raw or raw == '-':
                continue

            # 헤더 키워드 스킵
            if raw in SKIP_VALUES:
                continue

            # 서브 브랜드 헤더 감지
            sub_brand = is_sub_brand_header(raw)
            if sub_brand and not sub_brand_found:
                sub_brand_found = True
                # 서브 브랜드의 데이터를 이 행 이후로 처리
                self._process_block(
                    ws, oil_product, tier, sub_brand,
                    car_col, gas_col, diesel_col, dry_run,
                    start_row=row + 1, end_row=end_row,
                )
                break  # 현재 블록은 여기서 종료

            # 가격이 없는 메모 행 감지 (ex: "판촉서비스", "정기물 교체" 등)
            gas_val = ws.cell(row=row, column=gas_col + 1).value
            if gas_val is None and (diesel_col is None or ws.cell(row=row, column=diesel_col + 1).value is None):
                continue

            model_name, gen_name = parse_car_name(raw)
            if not model_name:
                continue

            car_model = self._get_or_create_model(brand, model_name, gen_name, dry_run)

            # 휘발유 가격
            gas_price = parse_price(ws.cell(row=row, column=gas_col + 1).value)
            if gas_price:
                self._save_price(car_model, oil_product, '휘발유', gas_price, dry_run)
                self._save_price(car_model, oil_product, '하이브리드', gas_price, dry_run)
                count += 1

            # 경유 가격
            if diesel_col is not None:
                diesel_price = parse_price(ws.cell(row=row, column=diesel_col + 1).value)
                if diesel_price:
                    self._save_price(car_model, oil_product, '경유', diesel_price, dry_run)
                    count += 1

        if count > 0:
            self.stdout.write(f'  {brand_name}: {count}건')

    def _get_or_create_brand(self, brand_name, dry_run):
        if brand_name in self.brand_cache:
            return self.brand_cache[brand_name]
        if dry_run:
            return None
        brand, created = CarBrand.objects.get_or_create(
            name=brand_name,
            defaults={'order': CarBrand.objects.count() + 1},
        )
        self.brand_cache[brand_name] = brand
        return brand

    def _get_or_create_model(self, brand, model_name, gen_name, dry_run):
        if dry_run or not brand:
            return None

        brand_id = brand.id

        if gen_name:
            # 부모 조회/생성
            parent_key = (brand_id, model_name, None)
            if parent_key not in self.model_cache:
                parent, created = CarModel.objects.get_or_create(
                    brand=brand, name=model_name, parent=None,
                    defaults={'order': 0},
                )
                self.model_cache[parent_key] = parent
                if created:
                    self.stats['models_created'] += 1

            parent = self.model_cache[parent_key]

            # 세대 조회/생성
            gen_key = (brand_id, gen_name, parent.id)
            if gen_key not in self.model_cache:
                gen, created = CarModel.objects.get_or_create(
                    brand=brand, name=gen_name, parent=parent,
                    defaults={'order': 0},
                )
                self.model_cache[gen_key] = gen
                if created:
                    self.stats['models_created'] += 1

            return self.model_cache[gen_key]
        else:
            key = (brand_id, model_name, None)
            if key not in self.model_cache:
                model, created = CarModel.objects.get_or_create(
                    brand=brand, name=model_name, parent=None,
                    defaults={'order': 0},
                )
                self.model_cache[key] = model
                if created:
                    self.stats['models_created'] += 1

            return self.model_cache[key]

    def _save_price(self, car_model, oil_product, fuel_name, price, dry_run):
        fuel_type = self.fuel_types.get(fuel_name)
        if not fuel_type or not car_model:
            self.stats['skipped'] += 1
            return
        if dry_run:
            self.stats['prices_created'] += 1
            return

        _, created = OilPrice.objects.update_or_create(
            car_model=car_model,
            oil_product=oil_product,
            fuel_type=fuel_type,
            defaults={'price': price},
        )
        if created:
            self.stats['prices_created'] += 1
        else:
            self.stats['prices_updated'] += 1
